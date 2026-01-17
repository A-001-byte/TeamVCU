from flask import Blueprint, request, jsonify
from flask_jwt_extended import jwt_required, get_jwt_identity
from ..extensions import db
from ..models.transaction import Transaction
from ..utils.categorizer import auto_categorize

import csv
from openpyxl import load_workbook

txn_bp = Blueprint("transactions", __name__)

# ------------------------
# Manual Transaction
# ------------------------
@txn_bp.route("/manual", methods=["POST"])
@jwt_required()
def add_transaction():
    user_id = get_jwt_identity()
    data = request.json

    txn = Transaction(
        user_id=user_id,
        amount=data["amount"],
        txn_type=data["txn_type"],
        merchant=data["merchant"],
        category=data.get("category") or auto_categorize(data["merchant"]),
        mode=data["mode"],
        source="MANUAL"
    )

    db.session.add(txn)
    db.session.commit()

    return jsonify({"message": "Transaction added"}), 201


# ------------------------
# Fetch Transactions
# ------------------------
@txn_bp.route("", methods=["GET"])
def get_transactions():
    # Use test user for now
    user_id = "test-user-123"
    txns = Transaction.query.filter_by(user_id=user_id).all()

    return jsonify([
        {
            "id": t.id,
            "amount": t.amount,
            "category": t.category,
            "merchant": t.merchant,
            "txn_type": t.txn_type,
            "date": t.txn_date.isoformat() if t.txn_date else None,
            "source": t.source
        }
        for t in txns
    ]), 200


# ------------------------
# CSV Upload
# ------------------------
@txn_bp.route("/upload", methods=["POST"])
def upload_transactions():
    # Generate a test user_id for now (testing without auth)
    user_id = "test-user-123"

    try:
        if "file" not in request.files:
            return {"error": "CSV file is required"}, 400

        file = request.files["file"]

        if file.filename == "":
            return {"error": "No selected file"}, 400

        if not file.filename.endswith(".csv"):
            return {"error": "Only CSV files allowed"}, 400

        stream = file.stream.read().decode("utf-8").splitlines()
        reader = csv.DictReader(stream)

        required_fields = {"amount", "merchant"}

        created = 0
        failed = 0

        for row in reader:
            if not required_fields.issubset(row.keys()):
                failed += 1
                continue

            try:
                merchant = row["merchant"]
                # Handle both "type" and "txn_type" field names
                txn_type = row.get("type") or row.get("txn_type") or "DEBIT"
                
                txn = Transaction(
                    user_id=user_id,
                    amount=float(row["amount"]),
                    txn_type=txn_type.upper(),
                    merchant=merchant,
                    category=row.get("category") or auto_categorize(merchant),
                    mode=row.get("mode", "UNKNOWN"),
                    source="CSV"
                )
                db.session.add(txn)
                created += 1
            except Exception as e:
                print(f"Error processing row: {e}")
                failed += 1

        db.session.commit()

        return {
            "message": "CSV processed",
            "created": created,
            "failed": failed
        }, 201
    except Exception as e:
        print(f"Upload error: {e}")
        return {"error": f"Upload failed: {str(e)}"}, 500


# ------------------------
# Excel Upload (.xlsx)
# ------------------------
@txn_bp.route("/upload-excel", methods=["POST"])
def upload_excel():
    # Use test user for now (testing without auth)
    user_id = "test-user-123"

    if "file" not in request.files:
        return {"error": "Excel file is required"}, 400

    file = request.files["file"]

    if file.filename == "":
        return {"error": "No selected file"}, 400

    if not file.filename.endswith(".xlsx"):
        return {"error": "Only .xlsx files are supported"}, 400

    try:
        workbook = load_workbook(file, data_only=True)
        sheet = workbook.active
    except Exception:
        return {"error": "Invalid Excel file"}, 400

    headers = [cell.value for cell in sheet[1]]
    required_fields = {"amount", "type", "merchant", "mode"}

    if not required_fields.issubset(set(headers)):
        return {
            "error": "Excel must contain headers: amount, type, merchant, mode"
        }, 400

    header_index = {header: idx for idx, header in enumerate(headers)}

    created = 0
    failed = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        try:
            amount = float(row[header_index["amount"]])
            txn_type = row[header_index["type"]].upper()
            merchant = row[header_index["merchant"]]
            mode = row[header_index["mode"]]

            category = (
                row[header_index["category"]]
                if "category" in header_index and row[header_index["category"]]
                else auto_categorize(merchant)
            )

            txn = Transaction(
                user_id=user_id,
                amount=amount,
                txn_type=txn_type,
                merchant=merchant,
                category=category,
                mode=mode,
                source="EXCEL"
            )

            db.session.add(txn)
            created += 1
        except Exception:
            failed += 1

    db.session.commit()

    return {
        "message": "Excel processed",
        "created": created,
        "failed": failed
    }, 201
